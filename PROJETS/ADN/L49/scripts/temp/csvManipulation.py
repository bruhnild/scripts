"""
    Date : 10/2017
    Author : g.bruel@metis-reseaux.fr

    Version 1.0
    
    This module allow user to :
        - create folder
        - paste files in this folder
        - paste contain files to a file
        - create pre-entry excel file
"""

import csv, os, shutil, time
import openpyxl
from openpyxl import load_workbook

# GLOBAL VARIABLE
# CHANGE IF NECESSARY

workingPath = "C:/Users/CitruX/Documents/scripts"
csvFileName = 'CSV.csv'
xlsxFileName = 'template.xlsx'
sheetsNames = ["Feuil1", "Feuil2"]
dateStr = time.strftime("%d%m%Y")

nameTemplate = "template.xlsx" # xls template location and name
locTemplate = "C://scripts"

folderLoc = ""
folderName = ""

# FUNCTION

def changeType(h): # change str to int if possible    
    try :        
        l = int(h)
        return l
    except ValueError:        
        return False

def callChange(word):
    n = ""
    if(changeType(word)):
        return changeType(word)
    else :
        for i in word :
            res = changeType(i)
            if (res):
                n = n + str(res)
                return int(n)

def createName (i, date):
    try :        
        name = str(i)+"_" + date        
        return name
    except ValueError:
        return False

def controlSheet (wbook,sheetString):
    """
        Return openpyxl sheet object or boolean
 
        :param wbook: openpyxl workbook
        :param sheetString: name of sheet to control
        :type wbook: object
        :type sheetString: string
        :return: sheet from workbook or boolean

        .. warning: only testing with xlsx files 
    """    
    try :
        # open sheet
        res=wbook[sheetString]
        return res
    except KeyError :
        return False

# FUNCTION
def pastToExcel(csvName, xlsFile, sheetToPaste):
    """
        paste only one sheet of csv file to a given excel file sheet
 
        :csvName: csv fileName
        :xlsFile:  xlsx file name
        :sheetToPaste: sheet name
        :type csvName: string
        :type xlsFile: string
        :type sheetToPaste: string
        :return: excel file saving with new values from csv

        .. warning: use only csv file with one sheet and test excel file if not .xlsx ! 
    """   
    data=[]    
    # open excel file
    wb=load_workbook(xlsFile)
    # return sheet if exist in excel file
    ws=controlSheet(wb, sheetToPaste)
    
    if ws != False :    # continue if sheet exist    
        # open csv file
        with open(csvName, 'r') as f:
            # read file
            reader=csv.reader(f, delimiter=';')
            # parse and past row to array
            for row in reader:
                data.append(row)

        # parse all lines
        line=1
        for r in data : # l param is line
            # lecture des colones
            col=1
            for p in r: # o param is column                    
                cellref=ws.cell(row=line, column=col)
                cellref.value=p
                col=col+1
            line=line+1
        # save modification in file
        wb.save(xlsFile)

# - STARTING CODE - #
os.chdir("C://scripts/"); # change work directory

# go to work directory
os.chdir(workingPath)
for sheets in sheetsNames:
    pastToExcel(csvFileName, xlsxFileName, sheets)
    
    
